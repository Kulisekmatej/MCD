"""Export souhrnu hodin do CSV a Excelu (.xlsx)."""

from __future__ import annotations

import csv
from typing import Iterable, List, Optional

from .models import EmployeeTotal

_HEADER = [
    "Zaměstnanec", "Úvazek", "Zletilost", "Hodiny (h:mm)", "Hodiny (des.)",
    "z toho víkend (h:mm)", "z toho víkend (des.)",
    "Noční (h:mm)", "Noční (des.)", "z toho noční víkend (h:mm)",
    "Dovolená (h:mm)", "Dovolená (des.)",
    "Nemocenská (h:mm)", "Nemocenská (des.)", "Počet směn", "Počet dní",
]


def _rows(totals: Iterable[EmployeeTotal]) -> List[list]:
    return [
        [t.employee, t.contract_label, t.status_label, t.hm, t.hours,
         t.weekend_hm, t.weekend_hours,
         t.night_hm, t.night_hours, t.weekend_night_hm,
         t.vacation_hm, t.vacation_hours, t.sick_hm, t.sick_hours,
         t.shift_count, t.day_count]
        for t in totals
    ]


def _hm(total_minutes: int) -> str:
    return f"{total_minutes // 60}:{total_minutes % 60:02d}"


def _total_row(totals: List[EmployeeTotal]) -> list:
    """Řádek „CELKEM" – musí mít stejné pořadí sloupců jako :data:`_HEADER`."""
    def s(attr):
        return sum(getattr(t, attr) for t in totals)
    tm, wk, ni, wn = s("total_minutes"), s("weekend_minutes"), s("night_minutes"), s("weekend_night_minutes")
    va, si = s("vacation_minutes"), s("sick_minutes")
    return [
        "CELKEM", "", "", _hm(tm), round(tm / 60.0, 2),
        _hm(wk), round(wk / 60.0, 2),
        _hm(ni), round(ni / 60.0, 2), _hm(wn),
        _hm(va), round(va / 60.0, 2), _hm(si), round(si / 60.0, 2),
        sum(t.shift_count for t in totals), "",
    ]


def export_csv(totals: Iterable[EmployeeTotal], path: str) -> None:
    totals = list(totals)
    rows = _rows(totals)
    with open(path, "w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.writer(fh, delimiter=";")
        writer.writerow(_HEADER)
        writer.writerows(rows)
        writer.writerow([])
        writer.writerow(_total_row(totals))


def export_xlsx(
    totals: Iterable[EmployeeTotal],
    path: str,
    period: Optional[str] = None,
) -> None:
    """Vytvoří .xlsx soubor. Vyžaduje knihovnu ``openpyxl``."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    totals = list(totals)
    wb = Workbook()
    ws = wb.active
    ws.title = "Hodiny"

    row_idx = 1
    if period:
        ws.cell(row=row_idx, column=1, value=f"Období: {period}").font = Font(bold=True)
        row_idx += 2

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="C8102E")  # McDonald's červená
    for col, title in enumerate(_HEADER, start=1):
        cell = ws.cell(row=row_idx, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    header_row = row_idx
    row_idx += 1

    for row in _rows(totals):
        for col, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col, value=value)
        row_idx += 1

    for col, value in enumerate(_total_row(totals), start=1):
        ws.cell(row=row_idx, column=col, value=value).font = Font(bold=True)

    # Šířky sloupců (musí odpovídat pořadí v _HEADER).
    widths = [28, 9, 11, 14, 14, 18, 18, 15, 15, 22, 16, 16, 17, 17, 12, 12]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + col)].width = width

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    wb.save(path)
